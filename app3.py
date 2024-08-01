import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Define file paths
file_paths = [
    r'C:\India Plants\Marketing\Shamil Report\Actual EXF Not Done.xlsx',
    r'C:\India Plants\Marketing\Shamil Report\Actual EXF Done.xlsx',
    # r'C:\India Plants\Marketing\Shamil Report\Actual WH Not Done.xlsx',
    r'C:\India Plants\Marketing\Shamil Report\Actual WH.xlsx',
    r'C:\India Plants\Marketing\Shamil Report\Actual EXF Not Done Upcoming.xlsx',
    r'C:\Users\srikanthve\OneDrive - Brandix Lanka Pvt Ltd\Desktop\sha_12.xlsx'
]

# Define sheet names
sheet_names = [
    'Actual EXF Not Done',
    'Actual EXF Done In Transit',
    # 'Actual WH Not Done',
    'Actual WH',
    'Actual EXF Not Done Upcoming',
    'Master Sheet'
]

# Columns to highlight
headers_to_highlight = [
    'Contracted ETD', 
    'ETA WH(Original Plan)', 
    'ETA WH(Revised Plan)', 
    'ETA WH (Actual)', 
    'Shipment Remark(Forwarder)', 
    'VDD or Not', 
    'EXF(Actual)',  
    'ETA WH_Condition',
    'Transportation Method(Actual)',
    'Ship to Port',
    'Item'
]

# Output file path
output_file_path = r'C:\India Plants\Marketing\Shamil Report\consolidated_report_new.xlsx'

# File uploader
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

if uploaded_file is not None:
    # Load the uploaded Excel file
    df = pd.read_excel(uploaded_file, sheet_name='Sheet1')
    
    # Convert relevant columns to datetime format
    date_columns = ['Contracted ETD', 'ETA WH(Original Plan)', 'ETA WH(Revised Plan)', 'ETA WH (Actual)']
    for col in date_columns:
        df[col] = pd.to_datetime(df[col])

    # Filter data based on scenarios
    today = datetime.now()
    start_date = today - timedelta(weeks=12)
    end_date = today + timedelta(days=21)

    # Scenario 1: Actual EXF Not Done
    scenario_1_df = df[(df['Contracted ETD'] >= start_date) & df['ETA WH (Actual)'].isna() & (df['EXF(Actual)'].isna())]
    scenario_1_df['ETA WH_Condition'] = (scenario_1_df['ETA WH(Original Plan)'] - scenario_1_df['ETA WH(Revised Plan)']).dt.days
    scenario_1_df = scenario_1_df[scenario_1_df['ETA WH_Condition'] < 0]

    # Scenario 2: Actual EXF Done
    scenario_2_df = df[(df['Contracted ETD'] >= start_date) & df['ETA WH (Actual)'].notna() & (df['EXF(Actual)'].notna())]
    scenario_2_df['ETA WH_Condition'] = (scenario_2_df['ETA WH(Original Plan)'] - scenario_2_df['ETA WH(Revised Plan)']).dt.days
    scenario_2_df = scenario_2_df[scenario_2_df['ETA WH_Condition'] < 0]

    # Scenario 4: Actual WH
    scenario_4_df = df[(df['Contracted ETD'] >= start_date) & df['ETA WH (Actual)'].notna()]
    scenario_4_df = scenario_4_df[scenario_4_df['ETA WH Delay Days'] > 0]

    # Scenario 5: Actual EXF Not Done Upcoming
    scenario_5_df = df[(df['Contracted ETD'] >= today) & (df['Contracted ETD'] <= end_date) & df['ETA WH (Actual)'].isna()]
    scenario_5_df['ETA WH_Condition'] = (scenario_5_df['ETA WH(Original Plan)'] - scenario_5_df['ETA WH(Revised Plan)']).dt.days
    scenario_5_df = scenario_5_df[scenario_5_df['ETA WH_Condition'] < 0]

    # Consolidation
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        scenario_1_df.to_excel(writer, sheet_name='Actual EXF Not Done', index=False)
        scenario_2_df.to_excel(writer, sheet_name='Actual EXF Done In Transit', index=False)
        # scenario_3_df.to_excel(writer, sheet_name='Actual WH Not Done', index=False)
        scenario_4_df.to_excel(writer, sheet_name='Actual WH', index=False)
        scenario_5_df.to_excel(writer, sheet_name='Actual EXF Not Done Upcoming', index=False)
        df.to_excel(writer, sheet_name='Master Sheet', index=False)

    # Apply highlighting
    wb = load_workbook(output_file_path)
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet_name in sheet_names:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.iter_cols(1, ws.max_column):
                if col[0].value in headers_to_highlight:
                    for cell in col:
                        cell.fill = highlight_fill

    wb.save(output_file_path)

    st.success(f"Data processed successfully! Consolidated report saved at: {output_file_path}")
    st.download_button(label="Download Consolidated Report", data=open(output_file_path, "rb").read(), file_name="consolidated_report_new.xlsx")
else:
    st.warning("Please upload an Excel file to process.")
